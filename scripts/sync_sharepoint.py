"""
LexCollect SharePoint Sync Script
Reads case lists and collector activity logs from Downloads folder,
detects new/updated records, and upserts them into Supabase.
Designed to run hourly via Windows Task Scheduler.
"""
import os, sys, json, hashlib, logging, configparser, zipfile
from datetime import datetime, date, time, timedelta
from pathlib import Path

import pandas as pd
from supabase import create_client
from openpyxl import load_workbook

# --- Config ---
SUPABASE_URL = "https://qbrufeewsisljtoegops.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InFicnVmZWV3c2lzbGp0b2Vnb3BzIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzMwNjY1MzAsImV4cCI6MjA4ODY0MjUzMH0.JI0cKKy_qDVzNBNsHlt4vDj0rZ2Kp7naNiCrvliNiD4"

DOWNLOADS = Path(os.path.expanduser("~/Downloads"))
SYNC_STATE_FILE = Path(__file__).parent / ".sync_state.json"
COLLECTOR_SYNC_SCHEMA_VERSION = "2026-04-09-reminders-v4"

ONEDRIVE_DOWNLOADS = Path(os.path.expanduser("~")) / "OneDrive - elizabethrosariolaw.com" / "Downloads"

CASE_LIST_FILES = {
    2022: DOWNLOADS / "2022 CASE LIST.xlsx",
    2023: ONEDRIVE_DOWNLOADS / "2023 CASE LIST - Copy.xlsx",
    2024: ONEDRIVE_DOWNLOADS / "2024 CASE LIST - Copy.xlsx",
    2025: DOWNLOADS / "2025 Case List (1).xlsx",
    2026: DOWNLOADS / "2026 Case List (2).xlsx",
}

COLLECTOR_LOG_FILENAMES = {
    "Alejandro A": "Alejandro A Collections Activity Log.xlsx",
    "Maritza V": "Maritza V Collections Activity Log.xlsx",
    "Patricio D": "Patricio D Collections Activity Log.xlsx",
}

COLLECTOR_LOG_FILES = {
    collector: DOWNLOADS / filename
    for collector, filename in COLLECTOR_LOG_FILENAMES.items()
}

COLLECTOR_LOG_SHORTCUTS = {
    collector: ONEDRIVE_DOWNLOADS.parent / f"{filename}.url"
    for collector, filename in COLLECTOR_LOG_FILENAMES.items()
}

MONTHLY_SHEETS = [
    "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
    "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER",
]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(Path(__file__).parent / "sync.log"),
    ],
)
log = logging.getLogger("sync")

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)


def load_sync_state():
    if SYNC_STATE_FILE.exists():
        return json.loads(SYNC_STATE_FILE.read_text())
    return {}


def save_sync_state(state):
    SYNC_STATE_FILE.write_text(json.dumps(state, indent=2))


def file_hash(path: Path) -> str:
    return hashlib.md5(path.read_bytes()).hexdigest()


def is_valid_xlsx(path: Path) -> bool:
    return path.exists() and path.suffix.lower() == ".xlsx" and zipfile.is_zipfile(path)


def get_shortcut_url(path: Path) -> str | None:
    if not path.exists():
        return None
    parser = configparser.RawConfigParser()
    try:
        parser.read(path, encoding="utf-8")
        return parser.get("InternetShortcut", "URL", fallback=None)
    except Exception:
        return None


def get_workbook_activity_max_date(path: Path) -> str | None:
    try:
        sheets = pd.read_excel(path, sheet_name=None, nrows=0)
        data_sheet_name = next(
            (
                name for name in sheets
                if "dropdown" not in name.lower()
                and "taxonomy" not in name.lower()
                and "csq" not in name.lower()
                and "call back" not in name.lower()
            ),
            None,
        )
        if not data_sheet_name:
            return None

        df = pd.read_excel(path, sheet_name=data_sheet_name, usecols=lambda c: str(c).strip().upper() == "DATE")
        if "DATE" not in df.columns:
            return None
        dates = pd.to_datetime(df["DATE"], errors="coerce").dropna()
        if dates.empty:
            return None
        return dates.max().strftime("%Y-%m-%d")
    except Exception as e:
        log.warning(f"Could not inspect activity max date for {path}: {e}")
        return None


def resolve_collector_log_file(collector: str) -> Path | None:
    filename = COLLECTOR_LOG_FILENAMES[collector]
    candidates = [
        DOWNLOADS / filename,
        ONEDRIVE_DOWNLOADS / filename,
        ONEDRIVE_DOWNLOADS.parent / filename,
    ]
    valid_candidates = [path for path in candidates if is_valid_xlsx(path)]

    shortcut = COLLECTOR_LOG_SHORTCUTS.get(collector)
    shortcut_url = get_shortcut_url(shortcut) if shortcut else None
    if shortcut_url:
        log.info(f"{collector} live SharePoint shortcut found: {shortcut_url}")

    if not valid_candidates:
        if shortcut_url:
            log.warning(
                f"{collector} has a SharePoint shortcut but no valid local .xlsx. "
                "Open/sync the file from SharePoint or download a fresh copy before sync can import it."
            )
        return None

    chosen = max(valid_candidates, key=lambda path: path.stat().st_mtime)
    max_activity_date = get_workbook_activity_max_date(chosen)
    log.info(
        f"{collector} using local workbook {chosen} "
        f"(modified {datetime.fromtimestamp(chosen.stat().st_mtime):%Y-%m-%d %H:%M:%S}, "
        f"max activity DATE {max_activity_date or 'unknown'})"
    )
    if shortcut_url and max_activity_date and max_activity_date < datetime.now().strftime("%Y-%m-%d"):
        log.warning(
            f"{collector} local workbook only has activity through {max_activity_date}. "
            "If SharePoint shows newer rows, the local file is stale and must be synced/downloaded."
        )
    return chosen


def safe_str(val):
    if pd.isna(val) or val is None:
        return None
    return str(val).strip() or None


def safe_date(val):
    if pd.isna(val) or val is None:
        return None
    try:
        if isinstance(val, (datetime, date)):
            return val.strftime("%Y-%m-%d")
        d = pd.to_datetime(val, errors="coerce")
        if pd.isna(d):
            return None
        return d.strftime("%Y-%m-%d")
    except:
        return None


def safe_float(val):
    if pd.isna(val) or val is None:
        return None
    try:
        return float(val)
    except:
        return None


def safe_int(val):
    if pd.isna(val) or val is None:
        return None
    try:
        return int(float(val))
    except:
        return None


def format_activity_label(activity_type: str | None, call_direction: str | None) -> str:
    activity = safe_str(activity_type)
    direction = safe_str(call_direction)
    if activity:
      normalized = activity.replace("_", " ").strip().title()
      if normalized in {"Outbound Call", "Inbound Call"}:
          return normalized
      return normalized
    if direction:
        return "Inbound call" if direction.lower() == "inbound" else "Outbound call"
    return "Call"


def format_outcome_label(outcome: str | None) -> str:
    value = safe_str(outcome)
    if not value:
        return None
    return value.replace("_", " ").strip().title()


def parse_time_string(value):
    if value is None:
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    raw = str(value).strip()
    if not raw:
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(raw, fmt).time()
        except ValueError:
            continue
    return None


def duration_to_excel_time(minutes):
    if minutes is None:
        return None
    try:
        total = int(float(minutes))
    except Exception:
        return None
    hours = total // 60
    mins = total % 60
    return time(hour=min(hours, 23), minute=min(mins, 59))


def normalize_log_key(collector, activity_date, client_name, activity, outcome, transaction_id):
    return "|".join([
        safe_str(collector) or "",
        safe_str(activity_date) or "",
        safe_str(client_name) or "",
        safe_str(activity) or "",
        safe_str(outcome) or "",
        safe_str(transaction_id) or "",
    ]).lower()


# =====================
# CASE LIST IMPORT
# =====================

def parse_monthly_intakes(filepath: Path, year: int) -> list[dict]:
    """Parse monthly intake sheets from a Case List workbook."""
    import warnings
    warnings.filterwarnings("ignore")

    try:
        sheets = pd.read_excel(filepath, sheet_name=None)
    except Exception as e:
        log.error(f"Failed to read {filepath}: {e}")
        return []

    month_map = {m: i + 1 for i, m in enumerate(MONTHLY_SHEETS)}
    # Build case-insensitive lookup: map uppercase month name → actual sheet name
    sheet_lookup = {}
    for actual_name in sheets.keys():
        sheet_lookup[actual_name.upper().strip()] = actual_name
    records = []

    for sheet_name in MONTHLY_SHEETS:
        actual_sheet = sheet_lookup.get(sheet_name)
        if actual_sheet is None:
            continue
        df = sheets[actual_sheet]
        if len(df) == 0 or "CLIENT" not in df.columns:
            continue

        for _, row in df.iterrows():
            client_name = safe_str(row.get("CLIENT"))
            if not client_name:
                continue

            case_number = safe_str(row.get("CASE #"))
            open_date = safe_date(row.get("DATE"))
            practice_area = safe_str(row.get("PRACTICE AREA"))
            lead_attorney = safe_str(row.get("LEAD ATTORNEY"))
            paralegal = safe_str(row.get("LEGAL ASSISTANT/PARALEGAL"))
            team = safe_str(row.get("TEAM"))

            # Normalize attorney name to title case
            if lead_attorney:
                lead_attorney = lead_attorney.strip().title()
            if paralegal:
                paralegal = paralegal.strip().title()

            records.append({
                "case_number": case_number,
                "case_name": client_name,
                "practice_area": normalize_practice_area(practice_area),
                "lead_attorney": lead_attorney,
                "paralegal": paralegal,
                "team": team,
                "open_date": open_date,
                "is_closed": False,
            })

    log.info(f"Parsed {len(records)} intake records from {filepath.name}")
    return records


def normalize_practice_area(raw: str | None) -> str | None:
    if not raw:
        return None
    r = raw.lower().strip()
    if r == "other" or r == "unknown":
        return None
    if "t visa" in r and "deriv" not in r:
        return "T Visa"
    if "t visa" in r and "deriv" in r:
        return "T Visa Derivative"
    if "t deriv" in r:
        return "T Visa Derivative"
    if ("vawa" in r or "360" in r) and ("marriage" in r or "marr" in r or " m" == r[-2:]):
        return "VAWA Marriage"
    if ("vawa" in r or "360" in r) and ("parent" in r or " p" == r[-2:]):
        return "VAWA Parent"
    if "vawa" in r or "360" in r:
        return "VAWA"
    if "u visa" in r or "u cert" in r or "u-visa" in r:
        return "U Visa / U Cert"
    if "foia" in r:
        return "FOIA"
    if "removal" in r or "nta" in r or "eoir" in r or "42-b" in r or "42b" in r:
        return "Removal Defense"
    if "i-130" in r or "130" in r:
        return "I-130 Family Petition"
    if "aos" in r or "adjustment" in r or "one step" in r or "245" in r:
        return "Adjustment of Status"
    if "i-765" in r or "ead" in r:
        return "I-765 EAD"
    if "u deriv" in r:
        return "U Visa / U Cert"
    if "601" in r or "waiver" in r:
        return "Other Immigration"
    if "military" in r or "daca" in r or "tps" in r or "n-400" in r or "n400" in r or "i-90" in r or "i-751" in r:
        return "Other Immigration"
    # Return raw value for anything unrecognized — let DB normalizer handle display
    return raw.strip()


def upsert_cases(records: list[dict]):
    """Upsert cases into immigration_cases by case_number."""
    if not records:
        return

    # Get existing case numbers
    existing = set()
    offset = 0
    while True:
        res = supabase.table("immigration_cases").select("case_number").range(offset, offset + 999).execute()
        if not res.data:
            break
        existing.update(r["case_number"] for r in res.data if r["case_number"])
        if len(res.data) < 1000:
            break
        offset += 1000

    new_records = [r for r in records if r.get("case_number") and r["case_number"] not in existing]

    if not new_records:
        log.info("No new cases to insert")
        return

    # Batch insert in chunks of 500
    for i in range(0, len(new_records), 500):
        batch = new_records[i:i + 500]
        try:
            supabase.table("immigration_cases").insert(batch).execute()
            log.info(f"Inserted {len(batch)} new cases (batch {i // 500 + 1})")
        except Exception as e:
            log.error(f"Failed to insert cases batch: {e}")


# =====================
# COLLECTOR LOG IMPORT
# =====================

def parse_collector_log(filepath: Path, collector_name: str) -> list[dict]:
    """Parse collector activity log."""
    import warnings
    warnings.filterwarnings("ignore")

    try:
        sheets = pd.read_excel(filepath, sheet_name=None)
    except Exception as e:
        log.error(f"Failed to read {filepath}: {e}")
        return []

    # First sheet is the data sheet (named after the collector)
    data_sheet = None
    for name, df in sheets.items():
        if "dropdown" not in name.lower() and "taxonomy" not in name.lower() and "csq" not in name.lower() and "call back" not in name.lower():
            data_sheet = df
            break

    if data_sheet is None or len(data_sheet) == 0:
        return []

    records = []
    for _, row in data_sheet.iterrows():
        client_name = safe_str(row.get("LEAD"))
        if not client_name:
            continue

        activity_date = safe_date(row.get("DATE"))
        if not activity_date:
            continue

        activity_type = safe_str(row.get("ACTIVITY"))
        outcome = safe_str(row.get("OUTCOME"))
        collected = safe_float(row.get("COLLECTED AMOUNT"))
        escalated_to = safe_str(row.get("ESCALATED TO"))
        origin = safe_str(row.get("ORIGIN"))
        start_time = safe_str(row.get("Start Time"))
        end_time = safe_str(row.get("End Time"))
        duration = safe_str(row.get("Activity Duration"))
        overdue_since = safe_date(row.get("OVERDUE SINCE"))
        delinquency_days = safe_int(row.get("DELINQUENCY"))
        next_payment = safe_str(row.get("NEXT PAYMENT"))
        case_status = safe_str(row.get("CASE STATUS"))
        notes = safe_str(row.get("NOTES")) or safe_str(row.get("TRANSACTION"))
        payment_id = safe_str(row.get("# PAYMENT"))
        commission = safe_float(row.get("COMMISSION"))

        # Map activity type to call_direction
        act_lower = (activity_type or "").lower()
        call_direction = "outbound"
        if "inbound" in act_lower:
            call_direction = "inbound"

        records.append({
            "collector": collector_name,
            "client_name": client_name,
            "activity_date": activity_date,
            "activity_type": act_lower.replace(" ", "_") or "call",
            "call_direction": call_direction,
            "outcome": outcome,
            "collected_amount": collected if collected and collected > 0 else None,
            "commission": commission,
            "escalated_to": escalated_to if escalated_to and escalated_to != "No need" else None,
            "origin": origin,
            "overdue_since": overdue_since,
            "delinquency_days": delinquency_days,
            "next_payment_expected": next_payment,
            "case_status": case_status,
            "notes": notes,
            "transaction_id": payment_id,
        })

    log.info(f"Parsed {len(records)} activity records from {filepath.name}")
    return records


def upsert_activities(records: list[dict]):
    """Insert new collector activities (deduplicated by collector+date+client)."""
    if not records:
        return

    # Get existing activities to avoid duplicates — check by unique date+collector combos
    existing_keys = set()
    existing_records = {}
    date_collector_pairs = set((r["activity_date"], r["collector"]) for r in records)
    for d, c in date_collector_pairs:
        try:
            res = supabase.table("collection_activities") \
                .select("id,collector,client_name,activity_date,next_payment_expected,case_status,overdue_since,delinquency_days,notes,escalated_to") \
                .eq("collector", c) \
                .eq("activity_date", d) \
                .limit(1000) \
                .execute()
            for r in (res.data or []):
                key = f"{r['collector']}|{r['activity_date']}|{r.get('client_name','')}"
                existing_keys.add(key)
                existing_records[key] = r
        except Exception as e:
            log.warning(f"Failed to check existing activities for {c} {d}: {e}")

    new_records = []
    update_count = 0
    for rec in records:
        key = f"{rec['collector']}|{rec['activity_date']}|{rec['client_name']}"
        if key not in existing_keys:
            new_records.append(rec)
            continue

        existing = existing_records.get(key)
        if not existing:
            continue

        updates = {}
        for field in ["next_payment_expected", "case_status", "overdue_since", "delinquency_days", "notes", "escalated_to"]:
            incoming = rec.get(field)
            if incoming is not None and incoming != "" and incoming != existing.get(field):
                updates[field] = incoming

        if updates:
            try:
                supabase.table("collection_activities").update(updates).eq("id", existing["id"]).execute()
                update_count += 1
            except Exception as e:
                log.warning(f"Failed to update existing activity {existing.get('id')}: {e}")

    if update_count:
        log.info(f"Updated {update_count} existing activity records with reminder/status fields")

    if not new_records:
        log.info("No new activity records to insert")
        return

    for i in range(0, len(new_records), 500):
        batch = new_records[i:i + 500]
        try:
            supabase.table("collection_activities").insert(batch).execute()
            log.info(f"Inserted {len(batch)} new activities (batch {i // 500 + 1})")
        except Exception as e:
            log.error(f"Failed to insert activities batch: {e}")


def get_collector_log_sheet_name(workbook):
    for name in workbook.sheetnames:
        lowered = name.lower()
        if "dropdown" not in lowered and "taxonomy" not in lowered and "csq" not in lowered and "call back" not in lowered:
            return name
    return None


def fetch_dashboard_activities_for_collector(collector: str) -> list[dict]:
    try:
        res = supabase.table("collection_activities") \
            .select("*") \
            .eq("collector", collector) \
            .order("activity_date", desc=False) \
            .limit(5000) \
            .execute()
        return res.data or []
    except Exception as e:
        log.error(f"Failed to fetch dashboard activities for {collector}: {e}")
        return []


def append_dashboard_activities_to_log(filepath: Path, collector: str):
    if not filepath.exists():
        log.warning(f"Collector log not found for export: {filepath}")
        return

    wb = load_workbook(filepath)
    sheet_name = get_collector_log_sheet_name(wb)
    if not sheet_name:
        log.warning(f"No data sheet found in {filepath.name}")
        return

    ws = wb[sheet_name]
    header_row = [cell.value for cell in ws[1]]
    header_index = {str(value).strip(): idx + 1 for idx, value in enumerate(header_row) if value}

    required_headers = [
        "ID", "WEEKDAY", "DATE", "COLLECTOR", "ACTIVITY", "LEAD", "ORIGIN", "OUTCOME",
        "ESCALATED TO", "COLLECTED AMOUNT", "Start Time", "End Time", "Activity Duration",
        "OVERDUE SINCE", "DELINQUENCY", "NEXT PAYMENT", "CASE STATUS", "COMMISSION",
        "# PAYMENT TRANSACTION", "NOTES",
    ]
    missing_headers = [h for h in required_headers if h not in header_index]
    if missing_headers:
        log.warning(f"{filepath.name} missing expected headers: {missing_headers}")
        return

    existing_keys = set()
    max_id = 0
    for row_idx in range(2, ws.max_row + 1):
        row_values = {header: ws.cell(row=row_idx, column=header_index[header]).value for header in required_headers}
        row_id = safe_int(row_values.get("ID"))
        if row_id and row_id > max_id:
            max_id = row_id
        existing_keys.add(normalize_log_key(
            row_values.get("COLLECTOR"),
            safe_date(row_values.get("DATE")),
            row_values.get("LEAD"),
            row_values.get("ACTIVITY"),
            row_values.get("OUTCOME"),
            row_values.get("# PAYMENT TRANSACTION"),
        ))

    remote_rows = fetch_dashboard_activities_for_collector(collector)
    append_count = 0

    for record in remote_rows:
        activity_date = safe_str(record.get("activity_date"))
        client_name = safe_str(record.get("client_name"))
        activity_label = format_activity_label(record.get("activity_type"), record.get("call_direction"))
        outcome_label = format_outcome_label(record.get("outcome"))
        transaction_id = safe_str(record.get("transaction_id"))
        dedupe_key = normalize_log_key(collector, activity_date, client_name, activity_label, outcome_label, transaction_id)
        if dedupe_key in existing_keys:
            continue

        max_id += 1
        row_idx = ws.max_row + 1
        row_data = {
            "ID": max_id,
            "WEEKDAY": safe_str(record.get("weekday")) or (pd.to_datetime(activity_date).strftime("%a") if activity_date else None),
            "DATE": pd.to_datetime(activity_date).to_pydatetime() if activity_date else None,
            "COLLECTOR": collector,
            "ACTIVITY": activity_label,
            "LEAD": client_name,
            "ORIGIN": safe_str(record.get("origin")) or "LexCollect",
            "OUTCOME": outcome_label,
            "ESCALATED TO": safe_str(record.get("escalated_to")) or "No need",
            "COLLECTED AMOUNT": safe_float(record.get("collected_amount")),
            "Start Time": parse_time_string(record.get("start_time")),
            "End Time": parse_time_string(record.get("end_time")),
            "Activity Duration": duration_to_excel_time(record.get("duration_minutes")),
            "OVERDUE SINCE": pd.to_datetime(record.get("overdue_since")).to_pydatetime() if record.get("overdue_since") else None,
            "DELINQUENCY": safe_int(record.get("delinquency_days")),
            "NEXT PAYMENT": safe_str(record.get("next_payment_expected")),
            "CASE STATUS": safe_str(record.get("case_status")),
            "COMMISSION": safe_float(record.get("commission")),
            "# PAYMENT TRANSACTION": transaction_id,
            "NOTES": safe_str(record.get("notes")),
        }

        for header, value in row_data.items():
            ws.cell(row=row_idx, column=header_index[header], value=value)

        existing_keys.add(dedupe_key)
        append_count += 1

    if append_count > 0:
        wb.save(filepath)
        log.info(f"Appended {append_count} dashboard activities to {filepath.name}")
    else:
        log.info(f"No dashboard activities needed export for {filepath.name}")


# =====================
# MAIN SYNC
# =====================

def run_sync():
    log.info("=" * 50)
    log.info("Starting LexCollect sync...")
    state = load_sync_state()

    # --- Case Lists ---
    for year, filepath in CASE_LIST_FILES.items():
        if not filepath.exists():
            log.warning(f"Case list not found: {filepath}")
            # Try alternate names
            alt = DOWNLOADS / f"{year} Case List.xlsx"
            if alt.exists():
                filepath = alt
            else:
                continue

        current_hash = file_hash(filepath)
        state_key = f"cases_{year}"
        if state.get(state_key) == current_hash:
            log.info(f"{year} case list unchanged, skipping")
            continue

        records = parse_monthly_intakes(filepath, year)
        upsert_cases(records)
        state[state_key] = current_hash

    # --- Collector Logs ---
    resolved_collector_files = {}
    for collector in COLLECTOR_LOG_FILENAMES:
        filepath = resolve_collector_log_file(collector)
        if filepath is None:
            continue
        resolved_collector_files[collector] = filepath
        if not filepath.exists():
            log.warning(f"Collector log not found: {filepath}")
            continue

        current_hash = file_hash(filepath)
        state_key = f"collector_{collector}_{COLLECTOR_SYNC_SCHEMA_VERSION}"
        if state.get(state_key) == current_hash:
            log.info(f"{collector} log unchanged, skipping")
            continue

        records = parse_collector_log(filepath, collector)
        upsert_activities(records)
        state[state_key] = current_hash

    # --- Reverse sync: dashboard activity -> collector logs ---
    for collector, filepath in resolved_collector_files.items():
        append_dashboard_activities_to_log(filepath, collector)

    save_sync_state(state)
    log.info("Sync complete!")


if __name__ == "__main__":
    try:
        run_sync()
    except Exception as e:
        log.exception(f"Sync failed: {e}")
        sys.exit(1)
