import json, re, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load Supabase contract data
with open(r'C:\Users\JeffreySoto-Gil\.claude\projects\C--Users-JeffreySoto-Gil-Apex\d6ff22bb-81f1-41f1-ab3e-6ccacb110b4d\tool-results\mcp-b42fba8b-bf11-4e9d-b287-e5771c1b8579-execute_sql-1777845284260.txt') as f:
    raw = f.read()
outer = json.loads(raw)
text = outer[0]['text']
result = json.loads(text)['result']
m = re.search(r'<untrusted-data[^>]+>\n(\[.*?\])\n</untrusted-data', result, re.DOTALL)
contracts = pd.DataFrame(json.loads(m.group(1)))
contracts['collected'] = pd.to_numeric(contracts['collected'], errors='coerce').fillna(0)
contracts['remaining'] = pd.to_numeric(contracts['remaining'], errors='coerce').fillna(0)
contracts['monthly_installment'] = pd.to_numeric(contracts['monthly_installment'], errors='coerce').fillna(0)

# Load Filevine April data
fv = pd.read_excel(r'C:\Users\JeffreySoto-Gil\Downloads\Invoice-Transaction Application 2026-04-28 1355.xlsx')
fv.columns = [c.strip() for c in fv.columns]
# Handle both serial-number and already-parsed datetime formats
if pd.api.types.is_numeric_dtype(fv['Transaction date']):
    base = pd.Timestamp('1899-12-30')
    fv['txn_date'] = base + pd.to_timedelta(fv['Transaction date'], unit='D')
    apr = fv[(fv['Transaction date'] >= 46115) & (fv['Transaction date'] <= 46140)].copy()
else:
    fv['txn_date'] = pd.to_datetime(fv['Transaction date'])
    fv['_serial'] = (fv['txn_date'] - pd.Timestamp('1899-12-30')).dt.days
    apr = fv[(fv['_serial'] >= 46115) & (fv['_serial'] <= 46140)].copy()
paid = apr[~apr['Transaction method'].str.lower().str.contains('write', na=False)].copy()

def extract_uuid(desc):
    if pd.isna(desc): return None
    m2 = re.search(r'payment ID: ([0-9a-f-]{36})', str(desc))
    return m2.group(1) if m2 else None

def extract_case(proj):
    m2 = re.search(r'\b(\d{2}-\d{4})\b', str(proj))
    return m2.group(1) if m2 else None

def extract_surname(proj):
    proj = str(proj).strip()
    if ',' in proj:
        s = proj.split(',')[0].strip()
        s = re.sub(r'\s*\([^)]*\)\s*$', '', s).strip()
        return s.upper()
    return proj.split()[0].upper() if proj else ''

paid['payment_uuid'] = paid['Transaction description'].apply(extract_uuid)
paid['case_num'] = paid['Project name'].apply(extract_case)
paid['surname'] = paid['Project name'].apply(extract_surname)

apr_summary = paid.groupby('Project name').agg(
    case_num=('case_num', 'first'),
    surname=('surname', 'first'),
    apr_paid=('Total amount of transaction', 'sum'),
    apr_txn_count=('Total amount of transaction', 'count'),
    last_txn_date=('txn_date', 'max'),
    last_txn_amount=('Total amount of transaction', 'last'),
    method=('Transaction method', lambda x: ', '.join(sorted(set(x.dropna()))))
).reset_index()

print(f"Filevine April clients: {len(apr_summary)}")
print(f"Supabase contracts: {len(contracts)}")

def best_contract(row, contracts_df):
    cn = row['case_num']
    sn = row['surname']
    if cn:
        hits = contracts_df[contracts_df['case_number'].fillna('').str.contains(cn.replace('-', ''), case=False)]
        if len(hits) == 1:
            return hits.iloc[0]
        if len(hits) > 0:
            for st in ['Active', 'Risk', 'Outstanding']:
                h = hits[hits['status'] == st]
                if len(h):
                    return h.iloc[0]
            return hits.iloc[0]
    if sn:
        hits = contracts_df[contracts_df['supabase_name'].str.upper().str.contains(sn, na=False)]
        if len(hits) == 1:
            return hits.iloc[0]
        if len(hits) > 0:
            for st in ['Active', 'Risk', 'Outstanding']:
                h = hits[hits['status'] == st]
                if len(h):
                    return h.iloc[0]
            return hits.iloc[0]
    return None

matched_rows = []
unmatched = []

for _, row in apr_summary.iterrows():
    contract = best_contract(row, contracts)
    if contract is not None:
        matched_rows.append({
            'filevine_project': row['Project name'],
            'case_num': row['case_num'] or '',
            'apr_paid': row['apr_paid'],
            'apr_txn_count': row['apr_txn_count'],
            'last_txn_date': row['last_txn_date'].date() if pd.notna(row['last_txn_date']) else '',
            'last_txn_amount': row['last_txn_amount'],
            'method': row['method'],
            'supabase_name': contract['supabase_name'],
            'client_id': contract['client_id'],
            'contract_id': contract['contract_id'],
            'contract_status': contract['status'],
            'delinquency_status': contract['delinquency_status'] or '',
            'sb_collected': contract['collected'],
            'sb_remaining': contract['remaining'],
            'sb_last_txn_date': contract['last_transaction_date'] or '',
            'sb_last_txn_amount': contract['last_transaction_amount'] or '',
            'next_due_date': contract['next_due_date'] or '',
            'monthly_installment': contract['monthly_installment'],
        })
    else:
        unmatched.append(row['Project name'])

df = pd.DataFrame(matched_rows)
print(f"Matched: {len(df)} | Unmatched: {len(unmatched)}")
if unmatched:
    print("Unmatched:")
    for u in unmatched:
        print(f"  {u}")

# Build Excel
wb = Workbook()
ws = wb.active
ws.title = "Filevine Clients Apr Review"

hdr_fill  = PatternFill('solid', start_color='1E3A5F')
risk_fill = PatternFill('solid', start_color='FFE0E0')
abn_fill  = PatternFill('solid', start_color='FFCCCC')
ok_fill   = PatternFill('solid', start_color='E2EFDA')
warn_fill = PatternFill('solid', start_color='FFF2CC')
thin = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

headers = [
    'Filevine Project', 'Case #', 'Apr Paid ($)', '# Txns', 'Last Txn Date',
    'Last Txn Amt', 'LexCollect Name', 'Contract Status', 'Delinquency',
    'LexCollect Collected', 'LexCollect Remaining', 'LexCollect Last Txn', 'Next Due',
    'Monthly Install.', 'Action Needed', 'Contract ID', 'Client ID'
]
ws.append(headers)
for col_idx in range(1, len(headers)+1):
    cell = ws.cell(1, col_idx)
    cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin
ws.row_dimensions[1].height = 28

status_order = {'Abandoned': 0, 'Risk': 1, 'Outstanding': 2, 'Active': 3}
delinq_order = {'Abandoned': 0, 'Delinquent': 1, 'Late': 2, 'Current': 3}
df['_so'] = df['contract_status'].map(lambda x: status_order.get(x, 9))
df['_do'] = df['delinquency_status'].map(lambda x: delinq_order.get(x, 9))
df = df.sort_values(['_so', '_do', 'filevine_project']).reset_index(drop=True)

for _, row in df.iterrows():
    st = row['contract_status']
    dq = row['delinquency_status']
    if st == 'Abandoned':
        action = 'ABANDONED - reopen or close in LexCollect'
    elif st == 'Risk' and dq == 'Delinquent':
        action = 'Update collected - paying via Filevine (Risk)'
    elif dq in ('Delinquent', 'Late'):
        action = 'Update collected - paying via Filevine'
    else:
        action = 'Verify - appears current'

    row_data = [
        row['filevine_project'],
        row['case_num'],
        row['apr_paid'],
        row['apr_txn_count'],
        str(row['last_txn_date']),
        row['last_txn_amount'],
        row['supabase_name'],
        st,
        dq,
        row['sb_collected'],
        row['sb_remaining'],
        str(row['sb_last_txn_date']),
        str(row['next_due_date']),
        row['monthly_installment'],
        action,
        row['contract_id'],
        row['client_id'],
    ]
    ws.append(row_data)
    r = ws.max_row

    if st == 'Abandoned':
        row_fill = abn_fill
    elif dq == 'Delinquent':
        row_fill = risk_fill
    elif dq == 'Late':
        row_fill = warn_fill
    else:
        row_fill = ok_fill

    for col_idx in range(1, len(headers)+1):
        cell = ws.cell(r, col_idx)
        cell.fill = row_fill
        cell.border = thin
        cell.font = Font(name='Arial', size=9)
        cell.alignment = Alignment(vertical='center')

    for ci in [3, 6, 10, 11, 14]:
        ws.cell(r, ci).number_format = '$#,##0.00'

col_widths = [40, 9, 13, 7, 13, 13, 32, 14, 12, 16, 16, 14, 12, 14, 38, 38, 38]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# Summary sheet
ws2 = wb.create_sheet('Summary')
n_abandoned = len(df[df['contract_status'] == 'Abandoned'])
n_delinquent = len(df[(df['contract_status'] == 'Risk') & (df['delinquency_status'] == 'Delinquent')])
n_late = len(df[df['delinquency_status'].isin(['Late', 'Delinquent']) & (df['contract_status'] == 'Active')])
n_current = len(df[df['delinquency_status'] == 'Current'])
summary_data = [
    ['LexCollect / Filevine Reconciliation - April 2026', ''],
    ['Generated', '2026-05-03'],
    ['', ''],
    ['TOTALS', ''],
    ['Filevine clients paying in April', len(df) + len(unmatched)],
    ['Matched to LexCollect contract', len(df)],
    ['No LexCollect contract found', len(unmatched)],
    ['', ''],
    ['BY STATUS', ''],
    ['Abandoned in LexCollect, still paying in Filevine', n_abandoned],
    ['Risk + Delinquent in LexCollect', n_delinquent],
    ['Active + Late/Delinquent in LexCollect', n_late],
    ['Appears current', n_current],
    ['', ''],
    ['FINANCIALS', ''],
    ['Total April Filevine collections (matched clients)', f"${df['apr_paid'].sum():,.2f}"],
    ['Total LexCollect remaining balance (these clients)', f"${df['sb_remaining'].sum():,.2f}"],
    ['', ''],
    ['COLOR KEY', ''],
    ['Red rows', 'Abandoned contract - immediate review'],
    ['Pink rows', 'Delinquent in LexCollect but paying in Filevine'],
    ['Yellow rows', 'Late in LexCollect but paying in Filevine'],
    ['Green rows', 'Current - verify sync is working'],
]
for s_row in summary_data:
    ws2.append(s_row)
ws2.column_dimensions['A'].width = 50
ws2.column_dimensions['B'].width = 22
for bold_row in [1, 4, 9, 15, 19]:
    ws2.cell(bold_row, 1).font = Font(bold=True, name='Arial', size=11)

output_path = r'C:\Users\JeffreySoto-Gil\Downloads\Filevine_Client_Review_Apr2026.xlsx'
wb.save(output_path)
print(f"\nSaved: {output_path}")
print(f"Main sheet rows: {len(df)}")
print(f"\nBreakdown:")
print(f"  Abandoned: {n_abandoned}")
print(f"  Risk+Delinquent: {n_delinquent}")
print(f"  Active+Late: {n_late}")
print(f"  Current: {n_current}")
print(f"  Unmatched: {len(unmatched)}")

# Also save the df as CSV for use by update script
df.to_csv(r'C:\Users\JeffreySoto-Gil\Downloads\filevine_matched_contracts.csv', index=False)
print("CSV saved for update script.")
